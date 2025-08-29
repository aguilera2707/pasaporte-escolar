from flask import (
    Flask, jsonify, request, render_template,
    redirect, url_for, session, flash, Response, send_file
)
from functools import wraps
from datetime import datetime, timedelta
from io import BytesIO, TextIOWrapper
import os
import csv
import unicodedata
import qrcode
from werkzeug.security import generate_password_hash
from app import app, db
from app.models import (
    Familia, MovimientoPuntos, Transaccion,
    Admin, Beneficio, LugarFrecuente,
    EventoQR, EventoQRRegistro
)
from app.models import LogEntry
from app import db
from flask import session
from datetime import datetime
import pytz

def registrar_log(action, entity, details=""):
    """Registra una acción en el Log de Actividades"""
    tz = pytz.timezone('America/Merida')
    entry = LogEntry(
        timestamp=datetime.utcnow(),
        user=session.get("nombre_usuario", "sistema"),
        role=session.get("rol", "desconocido"),
        action=action,
        entity=entity,
        details=details
    )
    db.session.add(entry)
    db.session.commit()

def login_requerido_admin(f):
    @wraps(f)
    def decorada(*args, **kwargs):
        if "admin_id" not in session:
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorada

@app.route("/admin")
@login_requerido_admin
def panel_admin():
    rol = session.get('rol')
    print(f"[DEBUG] Rol en sesión: {rol}")

    if "admin_id" not in session:
        return redirect(url_for("login"))

    eventos = EventoQR.query.order_by(EventoQR.id.desc()).all()
    familias = Familia.query.all()
    ultimo_evento = eventos[0] if eventos else None
    otros_eventos = eventos[1:] if len(eventos) > 1 else []

    return render_template(
        "admin.html",
        eventos=eventos,
        familias=familias,
        ultimo_evento=ultimo_evento,
        otros_eventos=otros_eventos,
        rol=rol  # IMPORTANTE: pasar el rol
    )

from flask import session, redirect, url_for, flash

# Decorador para roles
def requiere_rol(*roles_permitidos):
    def decorator(f):
        @wraps(f)
        def wrapper(*args, **kwargs):
            if 'admin_id' not in session:
                return redirect(url_for('login'))
            rol = session.get('rol')
            if rol not in roles_permitidos:
                flash("No tienes permiso para acceder a esta sección", "error")
                return redirect(url_for('panel_admin'))
            return f(*args, **kwargs)
        return wrapper
    return decorator


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        usuario = request.form["usuario"]
        password = request.form["password"]

        admin = Admin.query.filter_by(usuario=usuario).first()
        if admin and admin.verificar_password(password):
            session["admin_id"] = admin.id
            session["admin"] = admin.usuario
            session["rol"] = admin.rol
            session["nombre_usuario"] = admin.usuario  # <-- Guardamos el nombre del admin aquí
            print(f"[DEBUG] Login exitoso: {admin.usuario} con rol {admin.rol}")
            entry = LogEntry(
                timestamp=datetime.utcnow(),   
                user=admin.usuario,
                role=admin.rol,
                action="inicio sesión",
                entity="Admin",
                details="Inicio de sesión exitoso"
            )
            db.session.add(entry)
            db.session.commit()
            return redirect(url_for("panel_admin"))
        else:
            flash("Credenciales inválidas para administrador", "error")
            print("[DEBUG] Login fallido o usuario no encontrado")
            return render_template("login_unificado.html", active_tab="admin")

    # Método GET
    return render_template("login_unificado.html", active_tab="admin")


@app.route("/logout")
def logout():
    nombre = session.get("nombre_usuario", "<desconocido>")
    rol = session.get("rol", "Admin")
    session.clear()
    # --- Log de cierre de sesión ---
    entry = LogEntry(
        timestamp=datetime.utcnow(),   
        user=nombre,
        role=rol,
        action="cierre sesión",
        entity="Admin",
        details="El usuario cerró sesión"
    )
    db.session.add(entry)
    db.session.commit()
    return redirect(url_for("login"))


def quitar_acentos(cadena):
    return ''.join(
        c for c in unicodedata.normalize('NFD', cadena)
        if unicodedata.category(c) != 'Mn'
    )


@app.route("/lista_familias", methods=["POST"])
@login_requerido_admin
def lista_familias():
    import qrcode
    import os

    nombre = request.form.get("nombre")
    correo = request.form.get("correo")
    password = request.form.get("password")

    # Validar duplicado por correo
    if Familia.query.filter_by(correo=correo).first():
        flash("❌ Ese correo ya está registrado", "error")
        return redirect(url_for("panel_admin"))

    # Validar duplicado por nombre
    nombre_normalizado = quitar_acentos(nombre.strip().lower().replace(" ", ""))
    for fam in Familia.query.all():
        nombre_existente = quitar_acentos(fam.nombre.strip().lower().replace(" ", ""))
        if nombre_existente == nombre_normalizado:
            flash("❌ Ya existe una familia con ese nombre", "error")
            return redirect(url_for("panel_admin"))

    # Crear nueva familia con qr_version inicial
    nueva_familia = Familia(nombre=nombre, correo=correo, password=password, puntos=0)
    nueva_familia.qr_version = 1
    db.session.add(nueva_familia)
    db.session.commit()

    # Generar QR con versión en la URL
    qr_url = url_for("ver_familia", familia_id=nueva_familia.id, _external=True) + f"?version_qr={nueva_familia.qr_version}"
    qr = qrcode.make(qr_url)
    qr_path = os.path.join(app.root_path, "static", "qr", f"familia_{nueva_familia.id}.png")
    qr.save(qr_path)

    flash("✅ Familia creada exitosamente y QR generado", "success")
    return redirect(url_for("panel_admin"))


@app.route('/')
def index():
    return render_template('login_unificado.html')

from app.routes import quitar_acentos  # asegúrate de importar esta función si no está en el mismo archivo

@app.route("/familias", methods=["POST"])
def crear_familia():
    import qrcode
    import os

    data = request.get_json()
    nombre = data["nombre"]
    correo = data["correo"]
    password = data["password"]

    # ✅ Validar duplicado por nombre (ignorando mayúsculas, espacios y acentos)
    nombre_normalizado = quitar_acentos(nombre.strip().lower().replace(" ", ""))
    familias = Familia.query.all()
    for familia in familias:
        nombre_existente = quitar_acentos(familia.nombre.strip().lower().replace(" ", ""))
        if nombre_existente == nombre_normalizado:
            return jsonify({"error": "Ya existe una familia con ese nombre"}), 400

    # ✅ Validar duplicado por correo
    if Familia.query.filter_by(correo=correo).first():
        return jsonify({"error": "Ese correo ya está registrado"}), 400

    # ✅ Crear nueva familia
    nueva_familia = Familia(nombre=nombre, correo=correo, password=password)
    nueva_familia.qr_version = 1
    db.session.add(nueva_familia)    
    db.session.commit()

    # ✅ Generar código QR una sola vez
    qr_url = url_for('ver_familia', familia_id=nueva_familia.id, _external=True) + f"?version_qr={nueva_familia.qr_version}"
    qr = qrcode.make(qr_url)
    qr_path = os.path.join(app.root_path, 'static', 'qr', f'familia_{nueva_familia.id}.png')
    qr.save(qr_path)

    return jsonify({"mensaje": "Familia registrada exitosamente"}), 201



@app.route("/sumar_puntos", methods=["POST"])
def sumar_puntos():
    data = request.get_json()
    familia_id = data.get("familia_id")
    puntos = data.get("puntos")
    motivo = data.get("motivo", "Puntos sumados")

    familia = Familia.query.get(familia_id)
    if not familia:
        return jsonify({"error": "Familia no encontrada"}), 404

    familia.puntos += puntos
    movimiento = MovimientoPuntos(familia_id=familia.id, cambio=puntos, motivo=motivo)
    db.session.add(movimiento)
    db.session.commit()
    
    # 📝 Registrar en LOG
    registrar_log("suma", "Familia", f"Se sumaron {puntos} puntos a la familia '{familia.nombre}' (ID {familia.id}), motivo: {motivo}")
    
    # Enviar correo
    enviar_correo_movimiento(familia.correo, familia.nombre, puntos, "suma", motivo)

    return jsonify({"mensaje": "Puntos sumados correctamente", "puntos_actuales": familia.puntos})

@app.route("/restar_puntos", methods=["POST"])
@requiere_rol('admin', 'supervisor')
def restar_puntos():
    data = request.get_json()
    familia_id = data.get("familia_id")
    puntos = data.get("puntos")
    motivo = data.get("motivo", "Puntos restados")

    familia = Familia.query.get(familia_id)
    if not familia:
        return jsonify({"error": "Familia no encontrada"}), 404

    if familia.puntos < puntos:
        return jsonify({"error": "No hay suficientes puntos"}), 400

    familia.puntos -= puntos
    movimiento = MovimientoPuntos(familia_id=familia.id, cambio=-puntos, motivo=motivo)
    db.session.add(movimiento)
    db.session.commit()
    
        
    # 📝 Registrar en LOG
    registrar_log("resta", "Familia", f"Se restaron {puntos} puntos a la familia '{familia.nombre}' (ID {familia.id}), motivo: {motivo}")
    
    # Enviar correo
    enviar_correo_movimiento(familia.correo, familia.nombre, puntos, "canje", motivo)

    return jsonify({"mensaje": "Puntos restados correctamente", "puntos_actuales": familia.puntos})

@app.route("/historial/<int:familia_id>")
def historial(familia_id):
    movimientos = MovimientoPuntos.query.filter_by(familia_id=familia_id).order_by(MovimientoPuntos.fecha.desc()).all()
    datos = [
        {"cambio": m.cambio, "motivo": m.motivo, "fecha": m.fecha.strftime("%Y-%m-%d %H:%M")}
        for m in movimientos
    ]
    return jsonify(datos)


@app.route('/familias/<int:familia_id>/puntos', methods=['POST'])
def agregar_puntos(familia_id):
    data = request.get_json()

    if not data or 'puntos' not in data:
        return jsonify({'error': 'Se requiere el campo "puntos"'}), 400

    familia = Familia.query.get(familia_id)

    if not familia:
        return jsonify({'error': 'Familia no encontrada'}), 404

    try:
        puntos_a_agregar = int(data['puntos'])
        familia.puntos += puntos_a_agregar
        db.session.commit()
        return jsonify({'mensaje': f'Se han agregado {puntos_a_agregar} puntos a {familia.nombre}', 'puntos_totales': familia.puntos}), 200
    except ValueError:
        return jsonify({'error': 'El valor de "puntos" debe ser un número entero'}), 400
    
@app.route('/familias/<int:familia_id>/canjear', methods=['POST'])
def canjear_puntos(familia_id):
    data = request.get_json()

    if not data or 'puntos' not in data:
        return jsonify({'error': 'Se requiere el campo "puntos"'}), 400

    familia = Familia.query.get(familia_id)

    if not familia:
        return jsonify({'error': 'Familia no encontrada'}), 404

    try:
        puntos_a_restar = int(data['puntos'])

        if puntos_a_restar <= 0:
            return jsonify({'error': 'El número de puntos a canjear debe ser positivo'}), 400

        if familia.puntos < puntos_a_restar:
            return jsonify({'error': 'La familia no tiene suficientes puntos'}), 400

        familia.puntos -= puntos_a_restar
        db.session.commit()
        return jsonify({'mensaje': f'Se han canjeado {puntos_a_restar} puntos de {familia.nombre}', 'puntos_restantes': familia.puntos}), 200
    except ValueError:
        return jsonify({'error': 'El valor de "puntos" debe ser un número entero'}), 400
    
from datetime import datetime
import pytz

@app.route('/transaccion', methods=['POST'])
def registrar_transaccion():
    data = request.get_json()
    familia_id = data.get('familia_id')
    puntos = int(data.get('puntos'))
    tipo = data.get('tipo')
    descripcion = data.get('descripcion', '')

    familia = Familia.query.get(familia_id)
    if not familia:
        return jsonify({'error': 'Familia no encontrada'}), 404

    # Hora de Mérida
    hora_mexico = datetime.now(pytz.timezone('America/Mexico_City')).replace(tzinfo=None)

    if tipo == 'suma':
        familia.puntos += puntos

    elif tipo == 'canje':
        if familia.puntos < puntos:
            return jsonify({'error': 'Puntos insuficientes para canjear'}), 400
        familia.puntos -= puntos

    elif tipo == 'penalizacion':
        puntos = abs(puntos) * -1  # siempre restar
        familia.puntos += puntos  # restar puntos

        nueva_transaccion = Transaccion(
            familia_id=familia.id,
            tipo="penalizacion",
            puntos=puntos,
            descripcion=descripcion,
            fecha=hora_mexico
        )
        db.session.add(nueva_transaccion)
        db.session.add(familia)
        db.session.commit()

        # 📝 Registrar en LOG
        if tipo == "suma":
            registrar_log("suma", "Familia", f"Familia '{familia.nombre}' (ID {familia.id}) recibió {puntos} puntos. Motivo: {descripcion}")
        elif tipo == "canje":
            registrar_log("canje", "Familia", f"Familia '{familia.nombre}' (ID {familia.id}) canjeó {puntos} puntos. Motivo: {descripcion}")
        elif tipo == "penalizacion":
            registrar_log("penalizar", "Familia", f"Familia '{familia.nombre}' (ID {familia.id}) penalizada con {puntos} puntos. Motivo: {descripcion}")

        
        

        # Enviar correo con asunto de penalización
        enviar_correo_movimiento(
            familia.correo,
            "PENALIZACIÓN",
            f"Se ha aplicado una penalización de {abs(puntos)} puntos.\nMotivo: {descripcion or 'No especificado'}"
        )

        return jsonify({"mensaje": f"Se penalizó con {abs(puntos)} puntos."})

    else:
        return jsonify({'error': 'Tipo de transacción inválido'}), 400

    # --- Registro de historial para suma o canje ---
    nueva_transaccion = Transaccion(
        familia_id=familia_id,
        tipo=tipo,
        puntos=puntos if tipo == 'suma' else -puntos,
        descripcion=descripcion,
        fecha=hora_mexico
    )

    db.session.add(nueva_transaccion)
    db.session.commit()

    # Enviar correo automático para suma/canje
    enviar_correo_movimiento(familia.correo, familia.nombre, abs(puntos), tipo, descripcion)

    return jsonify({'mensaje': 'Transacción registrada con éxito', 'puntos_restantes': familia.puntos}), 200


    
@app.route('/familia/<int:familia_id>/historial', methods=['GET'])
def historial_transacciones(familia_id):
    familia = Familia.query.get(familia_id)
    if not familia:
        return jsonify({'error': 'Familia no encontrada'}), 404

    transacciones = Transaccion.query.filter_by(familia_id=familia_id).order_by(Transaccion.fecha.desc()).all()
    resultado = [
        {
            'id': t.id,
            'tipo': t.tipo,
            'puntos': t.puntos,
            'descripcion': t.descripcion,
            'fecha': t.fecha.strftime('%Y-%m-%d %H:%M:%S')
        }
        for t in transacciones
    ]

    return jsonify({'historial': resultado})

@app.route("/familias", methods=["GET"])
def obtener_familias():
    familias = Familia.query.all()
    resultado = []
    for familia in familias:
        resultado.append({
            "id": familia.id,
            "nombre": familia.nombre,
            "correo": familia.correo,
            "puntos": familia.puntos,
            "password": familia.password  
        })
    return jsonify(resultado)

from flask import render_template

@app.route('/familia/<int:familia_id>/ver')
def ver_familia(familia_id):
    familia = Familia.query.get(familia_id)
    if not familia:
        return "Familia no encontrada", 404

    # Obtener filtros desde la URL
    tipo = request.args.get('tipo')
    fecha_inicio = request.args.get('fecha_inicio')
    fecha_fin = request.args.get('fecha_fin')

    # Base de la consulta
    query = Transaccion.query.filter_by(familia_id=familia_id)

    # Filtro por tipo
    if tipo in ['suma', 'canje']:
        query = query.filter_by(tipo=tipo)

    # Filtro por fechas
    if fecha_inicio:
        try:
            fecha_inicio_dt = datetime.strptime(fecha_inicio, '%Y-%m-%d')
            query = query.filter(Transaccion.fecha >= fecha_inicio_dt)
        except ValueError:
            pass  # Ignora fechas inválidas

    if fecha_fin:
        try:
            fecha_fin_dt = datetime.strptime(fecha_fin, '%Y-%m-%d')
            fecha_fin_dt = fecha_fin_dt.replace(hour=23, minute=59, second=59)
            query = query.filter(Transaccion.fecha <= fecha_fin_dt)
        except ValueError:
            pass

    transacciones = query.order_by(Transaccion.fecha.desc()).all()

    # PASAR EL ROL ACTUAL A LA PLANTILLA
    rol_actual = session.get('rol', None)

    return render_template('familia.html', familia=familia, transacciones=transacciones, rol=rol_actual)


@app.route('/familia/<int:familia_id>/transaccion', methods=['POST'])
def registrar_transaccion_web(familia_id):
    data = request.get_json()
    tipo = data.get("tipo")
    try:
        puntos = int(data.get("puntos"))
    except (TypeError, ValueError):
        return jsonify({"error": "Puntos inválidos"}), 400

    descripcion = data.get("descripcion")
    familia = Familia.query.get(familia_id)
    if not familia:
        return jsonify({"error": "Familia no encontrada"}), 404

    if tipo not in ['suma', 'canje', 'penalizacion']:
        return jsonify({'error': 'Tipo de transacción no válido'}), 400


    if tipo == 'canje' and puntos > familia.puntos:
        return jsonify({"error": "No tienes suficientes puntos"}), 400

    puntos_finales = puntos if tipo == 'suma' else -abs(puntos)

    hora_mexico = datetime.now(pytz.timezone('America/Mexico_City')).replace(tzinfo=None)
    transaccion = Transaccion(
    familia_id=familia_id,
    tipo=tipo,
    puntos=puntos_finales,
    descripcion=descripcion,
    fecha=hora_mexico
)

    # **Este es el punto clave: ACTUALIZA EL CAMPO DE PUNTOS**
    familia.puntos += puntos_finales

    db.session.add(transaccion)
    db.session.commit()

    # 📝 Registrar en LOG
    if tipo == "suma":
        registrar_log("suma", "Familia", f"Familia '{familia.nombre}' (ID {familia.id}) recibió {puntos} puntos. Motivo: {descripcion}")
    elif tipo == "canje":
        registrar_log("canje", "Familia", f"Familia '{familia.nombre}' (ID {familia.id}) canjeó {puntos} puntos. Motivo: {descripcion}")
    elif tipo == "penalizacion":
        registrar_log("penalizar", "Familia", f"Familia '{familia.nombre}' (ID {familia.id}) penalizada con {puntos} puntos. Motivo: {descripcion}")

    
    
    enviar_correo_movimiento(familia.correo, familia.nombre, abs(puntos_finales), tipo, descripcion)

    return jsonify({"mensaje": "Transacción registrada con éxito"})




@app.route('/familia/<int:familia_id>/exportar')
def exportar_transacciones(familia_id):
    familia = Familia.query.get(familia_id)
    if not familia:
        return "Familia no encontrada", 404

    transacciones = Transaccion.query.filter_by(familia_id=familia_id).order_by(Transaccion.fecha.desc()).all()

    # Crear archivo CSV con codificación UTF-8 con BOM
    buffer = BytesIO()
    wrapper = TextIOWrapper(buffer, encoding='utf-8-sig', newline='')
    writer = csv.writer(wrapper)

    # Encabezados con columna de familia
    writer.writerow(['Familia', 'Fecha', 'Tipo', 'Puntos', 'Descripción'])

    for t in transacciones:
        writer.writerow([
            familia.nombre,
            t.fecha.strftime('%Y-%m-%d %H:%M'),
            'Acreditación' if t.tipo == 'suma' else 'Canje',
            t.puntos,
            t.descripcion or 'Sin descripción'
        ])

    wrapper.flush()
    buffer.seek(0)
    
    
    # 📝 Registrar en LOG
    registrar_log("exportar", "Transacciones", f"Exportación de transacciones de la familia '{familia.nombre}' (ID {familia.id})")


    return Response(
        buffer.read(),
        mimetype='text/csv',
        headers={
            "Content-Disposition": f"attachment; filename=transacciones_{familia.nombre}.csv"
        }
    )
    
@app.route('/crear_admin', methods=['GET', 'POST'])
def crear_admin():
    if 'admin_id' not in session:
        return redirect(url_for('login'))

    if request.method == 'POST':
        usuario = request.form.get('usuario')
        password = request.form.get('password')
        rol = request.form.get('rol')

        if not usuario or not password or not rol:
            flash("Todos los campos son obligatorios", "error")
            return render_template("crear_admin.html")

        if Admin.query.filter_by(usuario=usuario).first():
            flash("Ese usuario ya existe", "error")
            return render_template("crear_admin.html")

        nuevo_admin = Admin(usuario=usuario, rol=rol)
        nuevo_admin.set_password(password)  # ← Usa el método para guardar el hash

        db.session.add(nuevo_admin)
        db.session.commit()
        
        # 📝 Registrar en LOG
        registrar_log("crear", "Admin", f"Se creó administrador '{usuario}' con rol {rol}")

        flash("Usuario creado con éxito", "success")
        return redirect(url_for('crear_admin'))

    return render_template("crear_admin.html")



@app.route('/admin/usuarios')
@requiere_rol('admin')
def lista_administradores():
    if 'admin_id' not in session:
        return redirect(url_for('login'))   
    administradores = Admin.query.all()
    return render_template('lista_administradores.html', administradores=administradores) 

@app.route('/admin/usuarios/eliminar/<int:admin_id>', methods=['POST'])
@requiere_rol('admin')
def eliminar_administrador(admin_id):
    if 'admin_id' not in session:
        return redirect(url_for('login'))

    admin = Admin.query.get(admin_id)
    print("Intentando eliminar:", admin.usuario if admin else None)
    print("Logueado como:", session.get('admin'))

    # Validación 1: El usuario 'admin' no se puede eliminar nunca
    if not admin:
        flash("Administrador no encontrado", "error")
    elif admin.usuario.lower() == 'admin':
        flash("El usuario principal 'admin' no puede ser eliminado.", "error")
    # Validación 2: El admin logueado no puede eliminarse a sí mismo
    elif admin.usuario == session.get('admin'):
        flash("No puedes eliminar el usuario con el que estás logueado.", "error")
    else:
        db.session.delete(admin)
        db.session.commit()
        
        # 🗑 Admin eliminado
    entry = LogEntry(
        timestamp=datetime.utcnow(),
        user=session.get("nombre_usuario"),
        role=session.get("rol"),
        action="eliminar",
        entity="Administrador",
        details=f"id={admin_id}"
    )
    db.session.add(entry)
    db.session.commit()
        
    flash("Administrador eliminado exitosamente", "success")

    return redirect(url_for('lista_administradores'))



@app.route('/escanear')
def escanear_qr():
    if 'admin_id' not in session:
        return redirect(url_for('login'))
    return render_template('escanear.html')

# Mostrar la página lista_familias.html
@app.route("/lista_familias", methods=["GET"])
@login_requerido_admin
def mostrar_lista_familias():
    familias = Familia.query.all()
    return render_template("lista_familias.html", familias=familias)

@app.route("/familia/<int:familia_id>/qr")
def generar_qr_familia(familia_id):
    import qrcode
    import os

    familia = Familia.query.get_or_404(familia_id)

    # Si la familia no tiene versión aún, inicializar en 1
    if not familia.qr_version:
        familia.qr_version = 1
        db.session.commit()

    # Generar la URL con versionado
    qr_url = url_for("ver_familia", familia_id=familia.id, _external=True) + f"?version_qr={familia.qr_version}"
    qr = qrcode.make(qr_url)

    # Guardar el QR en static/qr/familia_<id>.png
    qr_path = os.path.join(app.root_path, "static", "qr", f"familia_{familia.id}.png")
    if os.path.exists(qr_path):
        try:
            os.remove(qr_path)  # limpiar si ya existía un QR viejo
        except OSError:
            pass
    qr.save(qr_path)

    return send_file(qr_path, mimetype="image/png")



from flask_login import login_user

@app.route("/login_familia", methods=["GET", "POST"])
def login_familia():
    if request.method == "POST":
        correo = request.form["correo"]
        password = request.form["password"]

        familia = Familia.query.filter_by(correo=correo).first()
        if familia and familia.password == password:  # 🔑 Comparación directa
            login_user(familia)
            return redirect(url_for("perfil_familia", familia_id=familia.id))
        else:
            flash("❌ Credenciales inválidas para familia", "error")
            return render_template("login_unificado.html", active_tab="familia")

    return render_template("login_unificado.html", active_tab="familia")


@app.route('/logout_familia')
def logout_familia():
    session.pop('familia_id', None)
    return redirect(url_for('login_familia'))

from flask_login import login_required, current_user
from flask import render_template

@app.route('/perfil_familia/<int:familia_id>')
@login_required
def perfil_familia(familia_id):
    # Si es administrador → puede ver cualquier familia
    if hasattr(current_user, "rol") and current_user.rol == "admin":
        familia = Familia.query.get_or_404(familia_id)
    else:
        # Si es familia → solo puede ver su propio perfil
        if current_user.id != familia_id:
            return "Acceso denegado", 403
        familia = current_user

    # Mantengo tu lógica previa
    transacciones = Transaccion.query.filter_by(familia_id=familia.id).order_by(Transaccion.fecha.desc()).all()
    beneficios = Beneficio.query.all()
    ultimo_evento = EventoQR.query.order_by(EventoQR.id.desc()).first()
    rol = session.get('rol')  # si aún usas el rol en los templates

    return render_template(
        'perfil_familia.html',
        familia=familia,
        transacciones=transacciones,
        beneficios=beneficios,
        ultimo_evento=ultimo_evento,
        rol=rol
    )




@app.route('/familia/<int:familia_id>/generar_qr', methods=['POST'])
def generar_o_regenerar_qr(familia_id):
    import os, qrcode
    from flask import flash, redirect, url_for
    from app import app, db

    familia = Familia.query.get_or_404(familia_id)

    # ⬆️ Incrementa versión del QR
    familia.qr_version = (familia.qr_version or 0) + 1
    db.session.commit()

    # 📦 Carpeta donde guardas el PNG del QR mostrado en familia.html
    qr_folder = os.path.join(app.root_path, 'static', 'qr')
    os.makedirs(qr_folder, exist_ok=True)

    # 🔗 Mantenemos tu ruta original, solo agregamos ?version_qr=N
    qr_url = url_for('ver_familia', familia_id=familia.id, _external=True) + f'?version_qr={familia.qr_version}'

    # 🖼️ Reescribe el archivo del QR con el nuevo contenido
    qr_img = qrcode.make(qr_url)
    qr_path = os.path.join(qr_folder, f'familia_{familia.id}.png')
    qr_img.save(qr_path)

    flash("Código QR actualizado correctamente", "success")
    # 👈 Te mantiene en el mismo template de la familia
    return redirect(url_for('ver_familia', familia_id=familia.id))


@app.route('/escanear_qr/<int:familia_id>', methods=['GET'])
def escanear_qr_suma(familia_id):
    familia = Familia.query.get(familia_id)
    if not familia:
        return "Familia no encontrada", 404

    puntos = 5  # Ajusta el número de puntos a sumar por escaneo
    familia.puntos += puntos

    movimiento = MovimientoPuntos(
        familia_id=familia.id,
        cambio=puntos,
        motivo="Puntos por escaneo QR"
    )
    db.session.add(movimiento)
    db.session.commit()

    return f"¡Se sumaron {puntos} puntos a {familia.nombre}!"

from flask import current_app


@app.route('/escanear_evento_directo/<int:evento_id>')
def escanear_evento_qr(evento_id):
    familia_id = session.get('familia_id')
    if not familia_id:
        return redirect(url_for('login_familia'))

    evento = EventoQR.query.get(evento_id)
    if not evento:
        return "Evento no encontrado", 404

    ya_escaneo = EventoQRRegistro.query.filter_by(familia_id=familia_id, evento_id=evento_id).first()
    if ya_escaneo:
        return render_template("asistencia_exitosa.html", evento=evento, familia=Familia.query.get(familia_id), ya_asistio=True)

    familia = Familia.query.get(familia_id)
    familia.puntos += evento.puntos

    db.session.add(EventoQRRegistro(familia_id=familia_id, evento_id=evento_id))
    db.session.add(Transaccion(
        familia_id=familia.id,
        tipo='suma',
        puntos=evento.puntos,
        descripcion=f"Asistencia al evento: {evento.nombre_evento}"
        
        
    ))
    
    
    db.session.commit()
    
    

    return render_template("asistencia_exitosa.html", evento=evento, familia=familia, ya_asistio=False)



@app.route('/admin/crear_beneficio', methods=['GET', 'POST'])
@requiere_rol('admin')
def crear_beneficio():
    if 'admin_id' not in session:
        return redirect(url_for('login'))

    # Obtener siempre los beneficios para mostrar la tabla
    beneficios = Beneficio.query.all()

    if request.method == 'POST':
        nombre = request.form.get('nombre')
        puntos_requeridos = request.form.get('puntos_requeridos')

        if not nombre or not puntos_requeridos:
            flash("Todos los campos son obligatorios", "error")
            return render_template('crear_beneficio.html', beneficios=beneficios)

        try:
            puntos_requeridos = int(puntos_requeridos)
        except ValueError:
            flash("Los puntos deben ser un número", "error")
            return render_template('crear_beneficio.html', beneficios=beneficios)

        beneficio = Beneficio(nombre=nombre, puntos_requeridos=puntos_requeridos)
        db.session.add(beneficio)
        db.session.commit()
        
        # 🎁 Beneficio creado
        entry = LogEntry(
            timestamp=datetime.utcnow(),
            user=session.get("nombre_usuario"),
            role=session.get("rol"),
            action="crear",
            entity="Beneficio",
            details=f"El administrador “{session.get('nombre_usuario')}” "f"creó el beneficio “{beneficio.nombre}” (ID {beneficio.id})."
        )
        db.session.add(entry)
        db.session.commit()

        flash("Beneficio creado exitosamente", "success")
        return redirect(url_for('crear_beneficio'))  # Redirige a la misma vista para ver el nuevo canje

    return render_template('crear_beneficio.html', beneficios=beneficios)


from datetime import datetime
import pytz

@app.route('/escanear_qr_evento/<int:evento_id>/<int:familia_id>')
def escanear_qr_familia_en_evento(evento_id, familia_id):
    evento = EventoQR.query.get(evento_id)
    if not evento:
        return "Evento no encontrado", 404

    familia = Familia.query.get(familia_id)
    if not familia:
        return "Familia no encontrada", 404

    ya_escaneo = EventoQRRegistro.query.filter_by(familia_id=familia_id, evento_id=evento_id).first()
    if ya_escaneo:
        return "Esta familia ya fue registrada en este evento", 400

    # Hora local de Mérida sin tzinfo
    hora_merida = datetime.now(pytz.timezone('America/Merida')).replace(tzinfo=None)

    # Sumar puntos
    familia.puntos += evento.puntos

    # Registrar asistencia
    db.session.add(EventoQRRegistro(
        familia_id=familia_id,
        evento_id=evento_id,
        fecha=hora_merida
    ))

    # Registrar transacción (para historial con hora)
    db.session.add(Transaccion(
        familia_id=familia_id,
        tipo='suma',
        puntos=evento.puntos,
        descripcion=f"Asistencia al evento: {evento.nombre_evento}",
        fecha=hora_merida
    ))

    # Enviar correo de confirmación
    enviar_correo_movimiento(
        correo=familia.correo,
        nombre_familia=familia.nombre,
        puntos=evento.puntos,
        tipo='suma',
        motivo=f"Asistencia al evento: {evento.nombre_evento}"
    )

    db.session.commit()

    return f"Puntos del evento '{evento.nombre_evento}' sumados correctamente a la familia '{familia.nombre}'."



@app.route('/staff/escanear/<int:familia_id>', methods=['GET', 'POST'])
@requiere_rol('admin', 'supervisor')
def escanear_familia_staff(familia_id):
    if 'admin_id' not in session:
        return redirect(url_for('login'))

    familia = Familia.query.get(familia_id)
    if not familia:
        return "Familia no encontrada", 404

    eventos = EventoQR.query.all()

    if request.method == 'POST':
        evento_id = int(request.form['evento_id'])

        ya_asistio = EventoQRRegistro.query.filter_by(
            familia_id=familia_id, evento_id=evento_id
        ).first()

        if ya_asistio:
            flash("Esta familia ya recibió puntos por este evento", "error")
        else:
            evento = EventoQR.query.get(evento_id)
            if not evento:
                flash("Evento no encontrado", "error")
                return redirect(url_for('escanear_familia_staff', familia_id=familia_id))

            # Actualiza puntos y registros
            familia.puntos += evento.puntos
            registro = EventoQRRegistro(familia_id=familia_id, evento_id=evento_id)
            db.session.add(registro)
            db.session.add(Transaccion(
                familia_id=familia_id,
                tipo='suma',
                puntos=evento.puntos,
                descripcion=f"Asistencia al evento: {evento.nombre_evento}"
            ))
            db.session.commit()
            
            # 📝 Registrar en LOG
            registrar_log(
                "escanear",
                "EventoQR",
                f"Staff ({session.get('nombre_usuario')}) registró asistencia de la familia '{familia.nombre}' (ID {familia.id}) al evento '{evento.nombre_evento}' (ID {evento.id})"
            )

            # 📧 Enviar correo de confirmación
            try:
                enviar_correo_movimiento(
                    familia.correo,
                    familia.nombre,
                    evento.puntos,
                    "suma",
                    f"Asistencia al evento: {evento.nombre_evento}"
                )
            except Exception as e:
                app.logger.error(f"Fallo al enviar correo (staff escaneo): {e}")

            flash(f"{evento.puntos} puntos sumados por el evento '{evento.nombre_evento}'", "success")

        return redirect(url_for('escanear_familia_staff', familia_id=familia_id))

    return render_template('escanear_staff.html', familia=familia, eventos=eventos)


@app.route('/familia/<int:familia_id>/generar_qr_staff', methods=['POST'])
def generar_qr_staff(familia_id):
    familia = Familia.query.get(familia_id)
    if not familia:
        return "Familia no encontrada", 404

    # Crear carpeta si no existe
    qr_folder = os.path.join('app', 'static', 'qr')
    os.makedirs(qr_folder, exist_ok=True)

    # Crear ruta de archivo
    qr_path = os.path.join(qr_folder, f'familia_{familia.id}.png')

    # ✅ Nueva URL para escaneo de staff
    qr_data = url_for('escanear_familia_staff', familia_id=familia.id, _external=True)
    qr = qrcode.make(qr_data)
    qr.save(qr_path)

    flash("Código QR actualizado para uso del staff", "success")
    return redirect(url_for('ver_familia', familia_id=familia.id))

@app.route('/admin/escanear_evento', methods=['GET'])
def escanear_evento_staff():
    if 'admin' not in session:
        return redirect(url_for('login'))

    eventos = EventoQR.query.order_by(EventoQR.id.desc()).all()
    return render_template('escanear_evento_staff.html', eventos=eventos)

# @app.route('/api/escanear_qr_evento', methods=['POST'])
# def escanear_qr_evento_api():
#     data = request.get_json()
#     evento_id = data.get('evento_id')
#     familia_id = data.get('familia_id')
#
#     if not evento_id or not familia_id:
#         return jsonify({"error": "Faltan datos"}), 400
#
#     evento = EventoQR.query.get(evento_id)
#     familia = Familia.query.get(familia_id)
#
#     if not evento:
#         return jsonify({"error": "Evento no encontrado"}), 404
#     if not familia:
#         return jsonify({"error": "Familia no encontrada"}), 404
#
#     # Verificar si ya escaneó este evento
#     ya_registrado = EventoQRRegistro.query.filter_by(evento_id=evento_id, familia_id=familia_id).first()
#     if ya_registrado:
#         return jsonify({"error": "Esta familia ya fue registrada en este evento"}), 409
#
#     # Registrar puntos y guardar asistencia
#     familia.puntos += evento.puntos
#     db.session.add(EventoQRRegistro(familia_id=familia_id, evento_id=evento_id))
#     db.session.add(Transaccion(
#         familia_id=familia_id,
#         tipo='suma',
#         puntos=evento.puntos,
#         descripcion=f"Asistencia al evento: {evento.nombre_evento}"
#     ))
#     db.session.commit()
#
#     return jsonify({
#         "mensaje": f"Puntos agregados correctamente por asistir al evento '{evento.nombre_evento}'",
#         "puntos_actuales": familia.puntos
#     }), 200


@app.route('/api/escanear_qr_evento', methods=['POST'])
def api_escanear_qr_evento():
    if 'admin_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401

    data = request.get_json()
    familia_id = data.get('familia_id')
    evento_id = data.get('evento_id')

    if not familia_id or not evento_id:
        return jsonify({'error': 'Datos incompletos'}), 400

    familia = Familia.query.get(familia_id)
    evento = EventoQR.query.get(evento_id)

    if not familia or not evento:
        return jsonify({'error': 'Familia o evento no encontrados'}), 404

    ya_escaneo = EventoQRRegistro.query.filter_by(
        familia_id=familia_id, evento_id=evento_id
    ).first()
    if ya_escaneo:
        return jsonify({'error': 'La familia ya escaneó este evento'}), 409

    # Actualiza puntos y guarda registros
    familia.puntos += evento.puntos
    db.session.add(EventoQRRegistro(familia_id=familia_id, evento_id=evento_id))
    db.session.add(Transaccion(
        familia_id=familia_id,
        tipo="suma",
        puntos=evento.puntos,
        descripcion=f"Asistencia al evento: {evento.nombre_evento}"
    ))
    db.session.commit()
    
    # 📝 Registrar en LOG
    registrar_log(
        "escanear",
        "EventoQR",
        f"Escaneo API → Familia '{familia.nombre}' (ID {familia.id}) asistió al evento '{evento.nombre_evento}' (ID {evento.id})"
    )

    # 📧 Enviar correo de confirmación
    try:
        enviar_correo_movimiento(
            familia.correo,
            familia.nombre,
            evento.puntos,
            "suma",
            f"Asistencia al evento: {evento.nombre_evento}"
        )
    except Exception as e:
        app.logger.error(f"Fallo al enviar correo (API escaneo): {e}")

    return jsonify({
        'mensaje': f"{evento.puntos} puntos asignados a {familia.nombre} por {evento.nombre_evento}",
        "puntos_actuales": familia.puntos
    }), 200


@app.route("/admin/escaneo_evento/<int:evento_id>")
def vista_escaneo_evento(evento_id):
    evento = EventoQR.query.get_or_404(evento_id)
    return render_template('escaneo_evento.html', evento=evento)


@app.route('/admin/registrar_asistencia_evento', methods=["POST"])
def registrar_asistencia_evento():
    from urllib.parse import urlparse, parse_qs
    import re
    import pytz
    from datetime import datetime

    data = request.get_json()
    qr_url = data.get("qr_url")
    evento_id = data.get("evento_id")

    if not qr_url or not evento_id:
        return jsonify({"error": "Datos incompletos"}), 400

    # Extraer familia_id del QR
    parsed = urlparse(qr_url)
    m = re.search(r'/familia/(\d+)', parsed.path)
    if not m:
        return jsonify({"error": "No se pudo extraer el ID de familia del QR"}), 400

    familia_id = int(m.group(1))
    familia = Familia.query.get(familia_id)
    evento = EventoQR.query.get(evento_id)

    if not familia or not evento:
        return jsonify({"error": "Evento o familia no encontrada"}), 404

    # Validar versión QR para evitar usar códigos antiguos
    qs = parse_qs(parsed.query)
    ver_param = qs.get('version_qr', [None])[0]
    try:
        version_capturada = int(ver_param) if ver_param is not None else None
    except ValueError:
        version_capturada = None

    vigente = familia.qr_version or 1
    if version_capturada is None or version_capturada != vigente:
        return jsonify({"error": "Este código QR está desactualizado. Usa el QR más reciente."}), 400

    # Evitar doble escaneo del mismo evento
    if EventoQRRegistro.query.filter_by(evento_id=evento.id, familia_id=familia.id).first():
        return jsonify({"error": f"{familia.nombre} ya registró asistencia a {evento.nombre_evento}"}), 400

    # Registrar asistencia y sumar puntos
    familia.puntos += evento.puntos
    db.session.add(EventoQRRegistro(evento_id=evento.id, familia_id=familia.id))
    db.session.add(Transaccion(
        familia_id=familia.id,
        tipo='suma',
        puntos=evento.puntos,
        descripcion=f"Asistencia al evento: {evento.nombre_evento}",
        fecha=datetime.now(pytz.timezone('America/Merida')).replace(tzinfo=None)
    ))
    db.session.commit()
    
    # 📝 Registrar en LOG
    registrar_log(
        "escanear",
        "EventoQR",
        f"Admin ({session.get('nombre_usuario')}) registró asistencia de la familia '{familia.nombre}' (ID {familia.id}) al evento '{evento.nombre_evento}' (ID {evento.id})"
    )
    
    

    # 💌 Enviar correo de confirmación
    try:
        enviar_correo_movimiento(
            destinatario=familia.correo,
            tipo='suma',
            puntos=evento.puntos,
            nombre_familia=familia.nombre,
            motivo=f"Asistencia al evento: {evento.nombre_evento}"
        )
    except Exception as e:
        app.logger.error(f"Fallo al enviar correo de movimiento: {e}")

    return jsonify({"mensaje": f"Asistencia registrada correctamente. Se sumaron {evento.puntos} puntos a {familia.nombre}"}), 200


from app.models import LugarFrecuente  # asegúrate de importarlo

from flask import request, render_template, redirect, url_for, flash, session, jsonify
from datetime import datetime
import pytz, os, qrcode
from geopy.distance import geodesic

from app import app, db
from app.models import EventoQR, EventoQRRegistro, Transaccion, LugarFrecuente, Familia

# --- Crear nuevo evento con QR ---
@app.route('/admin/crear_qr_evento', methods=['GET', 'POST'])
@requiere_rol('admin')
def crear_qr_evento():
    lugares = LugarFrecuente.query.all()
    eventos = EventoQR.query.order_by(EventoQR.id.desc()).all()

    if request.method == 'POST':
        nombre_evento = request.form['nombre_evento']
        puntos        = int(request.form['puntos'])
        latitud       = float(request.form['latitud'])
        longitud      = float(request.form['longitud'])
        requiere_ubic = 'requiere_ubic' in request.form
        valid_from    = request.form.get('valid_from')
        valid_to      = request.form.get('valid_to')

        evento = EventoQR(
            nombre_evento=nombre_evento,
            puntos=puntos,
            latitud=latitud,
            longitud=longitud,
            requiere_ubic=requiere_ubic,
            valid_from=datetime.strptime(valid_from, '%Y-%m-%dT%H:%M') if valid_from else None,
            valid_to=datetime.strptime(valid_to, '%Y-%m-%dT%H:%M')   if valid_to   else None
        )
        db.session.add(evento)
        db.session.flush()  # para obtener evento.id sin commit

        # Generar imagen QR
        qr_url   = url_for('escanear_evento_con_ubicacion', evento_id=evento.id, _external=True)
        qr_img   = qrcode.make(qr_url)
        qr_fname = f"evento_{evento.id}.png"
        qr_path  = os.path.join(app.root_path, 'static', 'qr_eventos', qr_fname)
        os.makedirs(os.path.dirname(qr_path), exist_ok=True)
        qr_img.save(qr_path)

        evento.qr_filename = qr_fname
        db.session.commit()
        
        # 📝 Registrar en LOG
        registrar_log("crear", "EventoQR", f"Evento '{nombre_evento}' creado con {puntos} puntos")

        flash('🎯 Evento QR creado con éxito.', 'success')
        return redirect(url_for('crear_qr_evento'))

    return render_template('crear_qr_evento.html', lugares=lugares, eventos=eventos)

from functools import wraps
from flask import session, redirect, url_for, flash

def login_requerido_familia(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if not session.get("familia_id"):
            flash("Debes iniciar sesión como familia para acceder.", "error")
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated_function

@app.route('/escanear_evento/<int:evento_id>')
@login_requerido_familia
def escanear_evento_con_ubicacion(evento_id):
    evento = EventoQR.query.get_or_404(evento_id)
    return render_template('escanear_evento_con_ubicacion.html', evento=evento)

from geopy.distance import geodesic


from datetime import datetime
import pytz
from geopy.distance import geodesic
from flask import request, jsonify, url_for

# … tus imports y decoradores …

@app.route('/validar_ubicacion_evento', methods=['POST'])
def validar_ubicacion_evento():
    data       = request.get_json()
    lat        = data.get("lat")
    lon        = data.get("lon")
    evento_id  = data.get("evento_id")
    familia_id = session.get("familia_id")

    if not evento_id or not familia_id:
        return jsonify({"error": "Datos incompletos"}), 400

    evento  = EventoQR.query.get(evento_id)
    familia = Familia.query.get(familia_id)
    if not evento or not familia:
        return jsonify({"error": "Evento o familia no encontrada"}), 404

    # --- 1) Validación rápida de rango de fecha/hora en hora local ---
    tz          = pytz.timezone("America/Mexico_City")
    ahora_local = datetime.now(tz)

    if evento.valid_from:
        vf_local = tz.localize(evento.valid_from)
        if ahora_local < vf_local:
            msg = vf_local.strftime('%d/%m/%Y %H:%M')
            return jsonify({
                "error": f"El evento aún no está activo. Disponible desde {msg}",
                "code": "fuera_de_fecha"
            }), 400

    if evento.valid_to:
        vt_local = tz.localize(evento.valid_to)
        if ahora_local > vt_local:
            msg = vt_local.strftime('%d/%m/%Y %H:%M')
            return jsonify({
                "error": f"El evento expiró. Válido hasta {msg}",
                "code": "qr_expirado"
            }), 400

    # --- 2) Validación de ubicación ---
    if evento.requiere_ubic:
        distancia = geodesic((evento.latitud, evento.longitud),
                            (float(lat), float(lon))).meters
        if distancia > 500:
            return jsonify({"redirect": url_for("ubicacion_invalida", evento_id=evento.id)})

    # --- 3) Escaneo duplicado ---
    if EventoQRRegistro.query.filter_by(evento_id=evento.id, familia_id=familia.id).first():
        return jsonify({"redirect": url_for("asistencia_exitosa", evento_id=evento.id, ya_asistio=1)})

    # --- 4) Registro exitoso ---
    familia.puntos += evento.puntos
    db.session.add(EventoQRRegistro(familia_id=familia.id, evento_id=evento.id))
    db.session.add(Transaccion(
        familia_id=familia.id,
        tipo='suma',
        puntos=evento.puntos,
        descripcion=f"Asistencia al evento: {evento.nombre_evento}"
    ))
    db.session.commit()

    print("[DEBUG] Llamando a enviar_correo_movimiento desde validar_ubicacion_evento")
    enviar_correo_movimiento(
        familia.correo,
        familia.nombre,
        evento.puntos,
        "suma",
        f"Asistencia al evento: {evento.nombre_evento}"
    )

    return jsonify({"redirect": url_for("asistencia_exitosa", evento_id=evento.id)})


@app.route("/asistencia_exitosa/<int:evento_id>")
def asistencia_exitosa(evento_id):
    ya_asistio = request.args.get("ya_asistio", default="0") == "1"
    evento = EventoQR.query.get_or_404(evento_id)
    familia_id = session.get("familia_id")
    familia = Familia.query.get_or_404(familia_id)
    return render_template("asistencia_exitosa.html", evento=evento, familia=familia, ya_asistio=ya_asistio)


@app.route('/ubicacion_invalida/<int:evento_id>')
def ubicacion_invalida(evento_id):
    evento = EventoQR.query.get_or_404(evento_id)
    return render_template('ubicacion_invalida.html', evento=evento)

@app.route('/admin/eliminar_evento/<int:evento_id>', methods=['POST'])
@requiere_rol('admin')
def eliminar_evento(evento_id):
    if 'admin_id' not in session:
        return redirect(url_for('login'))

    evento = EventoQR.query.get_or_404(evento_id)

    # Eliminar imagen QR si existe
    if evento.qr_filename:
        ruta_qr = os.path.join('app', 'static', 'qr_eventos', evento.qr_filename)
        if os.path.exists(ruta_qr):
            os.remove(ruta_qr)

    # Eliminar registros asociados (opcional si usas relaciones)
    registros = EventoQRRegistro.query.filter_by(evento_id=evento.id).all()
    for r in registros:
        db.session.delete(r)

    # Eliminar el evento
    db.session.delete(evento)
    db.session.commit()
    
    # 🗑 Evento eliminado
    entry = LogEntry(
        timestamp=datetime.utcnow(),
        user=session.get("nombre_usuario"),
        role=session.get("rol"),
        action="eliminar",
        entity="Evento",
        details=f"id={evento_id}"
    )
    db.session.add(entry)
    db.session.commit()

    flash("Evento eliminado correctamente", "exito")
    return redirect(url_for('panel_admin'))

from flask import send_from_directory

@app.route('/descargar_qr_evento/<filename>')
def descargar_qr_evento(filename):
    carpeta_qr = os.path.join(app.root_path, 'static', 'qr_eventos')
    
    # 📝 Registrar en LOG
    registrar_log("exportar", "EventoQR", f"Se descargó QR del evento (archivo {filename})")

    
    return send_from_directory(carpeta_qr, filename, as_attachment=True)

@app.route('/descargar_qr/<int:familia_id>')
def descargar_qr(familia_id):
    ruta = os.path.join(app.root_path, 'static', 'qr', f'familia_{familia_id}.png')
    return send_file(ruta, as_attachment=True)

#@app.route('/escanear_evento/<int:evento_id>')
#def escanear_evento_directo(evento_id):
#   if "familia_id" not in session:
#        return redirect(url_for("login_familia"))
#
#    evento = EventoQR.query.get_or_404(evento_id)
#    return render_template("escanear_evento.html", evento=evento)

@app.route('/familia/escanear_evento')
def escanear_evento_desde_familia():
    if 'familia_id' not in session:
        return redirect(url_for('login_familia'))
    return render_template('escanear_evento_desde_familia.html')


from datetime import datetime
import pytz
from flask import request, jsonify, url_for, session
from geopy.distance import geodesic
from app import db
from app.models import EventoQR, Familia, EventoQRRegistro, Transaccion




# Ruta en routes.py
from flask import render_template, request, redirect, url_for, flash
from app.models import Familia
from app import db

@app.route('/familia/<int:familia_id>/editar', methods=['GET', 'POST'])
def editar_familia(familia_id):
    familia = Familia.query.get_or_404(familia_id)
    if request.method == 'POST':
        familia.nombre = request.form['nombre']
        familia.correo = request.form['correo']
        familia.password = request.form['password']
        db.session.commit()

        # 📝 Familia editada
        entry = LogEntry(
            timestamp=datetime.utcnow(),
            user=session.get("nombre_usuario"),
            role=session.get("rol"),
            action="editar",
            entity="Familia",
            details=f"id={familia.id}, nombre={familia.nombre}"
        )
        db.session.add(entry)
        db.session.commit()

        flash('Familia actualizada correctamente', 'success')
        return render_template('editar_familia.html', familia=familia)
    return render_template('editar_familia.html', familia=familia)




@app.route('/admin/crear_lugar', methods=['GET', 'POST'])
def crear_lugar():
    if 'admin_id' not in session:
        return redirect(url_for('login'))

    if request.method == 'POST':
        nombre = request.form['nombre']
        latitud = request.form['latitud']
        longitud = request.form['longitud']
        nuevo_lugar = LugarFrecuente(nombre=nombre, latitud=latitud, longitud=longitud)
        db.session.add(nuevo_lugar)
        db.session.commit()
        flash('Lugar creado con éxito.', 'success')
        return redirect(url_for('crear_lugar'))

    lugares = LugarFrecuente.query.all()
    return render_template('crear_lugar.html', lugares=lugares)

# Eliminar lugar
@app.route('/admin/eliminar_lugar/<int:id>', methods=['POST'])
def eliminar_lugar(id):
    if 'admin_id' not in session:
        return redirect(url_for('login'))

    lugar = LugarFrecuente.query.get_or_404(id)
    db.session.delete(lugar)
    db.session.commit()
    flash('Lugar eliminado correctamente.', 'success')
    return redirect(url_for('crear_lugar'))

# Editar lugar
@app.route('/admin/editar_lugar/<int:id>', methods=['GET', 'POST'])
def editar_lugar(id):
    if 'admin_id' not in session:
        return redirect(url_for('login'))

    lugar = LugarFrecuente.query.get_or_404(id)

    if request.method == 'POST':
        lugar.nombre = request.form['nombre']
        lugar.latitud = request.form['latitud']
        lugar.longitud = request.form['longitud']
        db.session.commit()
        flash('Lugar actualizado correctamente.', 'success')
        return redirect(url_for('crear_lugar'))

    return render_template('editar_lugar.html', lugar=lugar)

@app.route('/admin/eliminar_eventos', methods=['POST'])
def eliminar_eventos():
    if 'admin_id' not in session:
        return redirect(url_for('login'))

    ids = request.form.getlist('eventos_seleccionados')

    for evento_id in ids:
        # Eliminar registros relacionados
        registros = EventoQRRegistro.query.filter_by(evento_id=evento_id).all()
        for r in registros:
            db.session.delete(r)

        # Eliminar evento
        evento = EventoQR.query.get(evento_id)
        if evento:
            # Borrar el archivo QR si existe
            ruta_qr = os.path.join('app', 'static', 'qr_eventos', evento.qr_filename)
            if os.path.exists(ruta_qr):
                os.remove(ruta_qr)

            db.session.delete(evento)

    db.session.commit()
    
    # 📝 Registrar en LOG
    registrar_log("eliminar", "EventoQR", f"Se eliminaron {len(ids)} eventos masivamente")
    flash('Eventos eliminados correctamente', 'success')
    return redirect(url_for('crear_qr_evento'))

from flask import render_template, request, redirect, url_for, flash, session
from app.models import EventoQR, LugarFrecuente, EventoQRRegistro, Transaccion
from app import db
from functools import wraps
from datetime import datetime

# ... tu decorador requiere_rol ...

@app.route('/admin/editar_evento/<int:evento_id>', methods=['GET', 'POST'])
@requiere_rol('admin')
def editar_evento(evento_id):
    evento = EventoQR.query.get_or_404(evento_id)
    lugares = LugarFrecuente.query.all()   # <-- cargamos los lugares frecuentes

    if request.method == 'POST':
        evento.nombre_evento = request.form['nombre_evento']
        evento.puntos         = int(request.form['puntos'])
        evento.latitud        = float(request.form['latitud'])
        evento.longitud       = float(request.form['longitud'])
        evento.requiere_ubic  = 'requiere_ubic' in request.form

        valid_from = request.form.get('valid_from')
        valid_to   = request.form.get('valid_to')

        evento.valid_from = datetime.strptime(valid_from, '%Y-%m-%dT%H:%M') if valid_from else None
        evento.valid_to   = datetime.strptime(valid_to,   '%Y-%m-%dT%H:%M') if valid_to   else None

        db.session.commit()
        
        # 📝 Registrar en LOG
        registrar_log("editar", "EventoQR", f"Evento ID {evento.id} actualizado: {evento.nombre_evento}")
        flash('Evento actualizado correctamente.', 'success')
        return redirect(url_for('crear_qr_evento'))

    return render_template('editar_evento.html', evento=evento, lugares=lugares)




@app.route('/admin/editar_beneficio/<int:beneficio_id>', methods=['GET', 'POST'])
def editar_beneficio(beneficio_id):
    if 'admin_id' not in session:
        return redirect(url_for('login'))

    beneficio = Beneficio.query.get_or_404(beneficio_id)

    if request.method == 'POST':
        nombre = request.form.get('nombre')
        puntos_requeridos = request.form.get('puntos_requeridos')

        if not nombre or not puntos_requeridos:
            flash("Todos los campos son obligatorios", "error")
            return redirect(url_for('editar_beneficio', beneficio_id=beneficio.id))

        try:
            puntos_requeridos = int(puntos_requeridos)
        except ValueError:
            flash("Los puntos deben ser un número", "error")
            return redirect(url_for('editar_beneficio', beneficio_id=beneficio.id))

        beneficio.nombre = nombre
        beneficio.puntos_requeridos = puntos_requeridos
        db.session.commit()
        # 📝 Registrar en LOG
        registrar_log("editar", "Beneficio", f"Beneficio ID {beneficio.id} actualizado: {beneficio.nombre}")
        flash("Beneficio actualizado exitosamente", "success")
        return redirect(url_for('crear_beneficio'))

    return render_template('editar_beneficio.html', beneficio=beneficio)

@app.route('/admin/eliminar_beneficio/<int:beneficio_id>', methods=['POST'])
@requiere_rol('admin')
def eliminar_beneficio(beneficio_id):
    if 'admin_id' not in session:
        return redirect(url_for('login'))

    beneficio = Beneficio.query.get_or_404(beneficio_id)
    db.session.delete(beneficio)
    db.session.commit()
    
        # 🗑 Beneficio eliminado
    entry = LogEntry(
        timestamp=datetime.utcnow(),
        user=session.get("nombre_usuario"),
        role=session.get("rol"),
        action="eliminar",
        entity="Beneficio",
        details=f"id={beneficio_id}"
    )
    db.session.add(entry)
    db.session.commit()
    flash("Beneficio eliminado exitosamente", "success")
    return redirect(url_for('crear_beneficio'))

from flask import send_file
import csv
from io import BytesIO, TextIOWrapper

# Vista para mostrar historial de escaneos de un evento
@app.route('/admin/evento/<int:evento_id>/historial')
def historial_escaneos_evento(evento_id):
    if 'admin_id' not in session:
        return redirect(url_for('login'))

    evento = EventoQR.query.get_or_404(evento_id)
    registros = EventoQRRegistro.query.filter_by(evento_id=evento_id).order_by(EventoQRRegistro.id.desc()).all()
    # Cargar familia con .familia gracias a la relación
    return render_template('historial_escaneos_evento.html', evento=evento, registros=registros)

@app.route('/admin/evento/<int:evento_id>/exportar')
def exportar_historial_escaneos_evento(evento_id):
    evento = EventoQR.query.get_or_404(evento_id)
    registros = EventoQRRegistro.query.filter_by(evento_id=evento_id).order_by(EventoQRRegistro.fecha.desc()).all()

    buffer = BytesIO()
    # Envolvemos el buffer para escribir texto
    wrapper = TextIOWrapper(buffer, encoding='utf-8-sig', newline='', write_through=True)
    writer = csv.writer(wrapper)

    writer.writerow(['ID Familia', 'Nombre Familia', 'Fecha de escaneo'])

    for r in registros:
        nombre_familia = r.familia.nombre if r.familia else "-"
        fecha_local = (r.fecha - timedelta(hours=6)).strftime('%Y-%m-%d %H:%M')
        writer.writerow([r.familia_id, nombre_familia, fecha_local])

    # No cerramos wrapper, solo hacemos flush
    wrapper.flush()
    buffer.seek(0)  # Volvemos al inicio para leer todo el contenido
    
    # 📝 Registrar en LOG
    registrar_log("exportar", "EventoQR", f"Se exportó historial del evento '{evento.nombre_evento}' (ID {evento.id})")


    return Response(
        buffer.read(),
        mimetype='text/csv',
        headers={
            "Content-Disposition": f"attachment; filename=historial_escaneos_evento_{evento.id}.csv"
        }
    )


#GENERACION DE PDF CON QR

from io import BytesIO
from flask import make_response
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
import qrcode
import os

@app.route('/familia/<int:familia_id>/descargar_pdf_qr')
def descargar_pdf_qr_familia(familia_id):
    familia = Familia.query.get_or_404(familia_id)

    # Generar URL que irá en el QR
    url = url_for('ver_familia', familia_id=familia.id, _external=True)

    # Crear QR
    qr_img = qrcode.make(url)
    qr_bytes = BytesIO()
    qr_img.save(qr_bytes)
    qr_bytes.seek(0)

    # Crear PDF en memoria
    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=letter)

    # Título: Nombre de la familia
    c.setFont("Helvetica-Bold", 18)
    c.drawCentredString(4.25 * inch, 10 * inch, f"QR de la familia: {familia.nombre}")

    # Insertar QR en PDF
    qr_temp_path = os.path.join('app', 'static', f'temp_qr_{familia.id}.png')
    qr_img.save(qr_temp_path)
    c.drawImage(qr_temp_path, 2.5 * inch, 6 * inch, width=3*inch, height=3*inch)

    # Finalizar PDF
    c.showPage()
    c.save()
    buffer.seek(0)

    # Eliminar archivo temporal QR
    os.remove(qr_temp_path)

    # Preparar respuesta con PDF
    response = make_response(buffer.getvalue())
    response.headers.set('Content-Type', 'application/pdf')
    response.headers.set('Content-Disposition', 'attachment', filename=f'qr_familia_{familia.id}.pdf')

    # 📝 Registrar en LOG
    registrar_log("exportar", "Familia", f"Se descargó PDF con QR de la familia '{familia.nombre}' (ID {familia.id})")


    return response

from app.models import EventoQR  # Asegúrate de importar tu modelo
from flask import abort

@app.route('/evento/<int:evento_id>/descargar_pdf_qr')
def descargar_pdf_qr_evento(evento_id):
    evento = EventoQR.query.get(evento_id)
    if not evento:
        abort(404, description="Evento no encontrado")

    # Ruta completa del QR guardado en static
    qr_path = os.path.join('app', 'static', 'qr_eventos', evento.qr_filename)

    if not os.path.exists(qr_path):
        abort(404, description="QR no encontrado")

    # Crear PDF en memoria
    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=letter)
    width, height = letter

    # Título
    c.setFont("Helvetica-Bold", 20)
    c.drawCentredString(width / 2, height - 100, evento.nombre_evento)

    # Insertar imagen QR (centrada)
    qr_size = 200
    c.drawImage(qr_path, (width - qr_size) / 2, height - 350, qr_size, qr_size)

    c.showPage()
    c.save()

    buffer.seek(0)
    
    # 📝 Registrar en LOG
    registrar_log("exportar", "EventoQR", f"Se descargó PDF con QR del evento '{evento.nombre_evento}' (ID {evento.id})")

    
    return send_file(buffer, as_attachment=True, download_name=f"{evento.nombre_evento}_QR.pdf", mimetype='application/pdf')

@app.route('/puntos-masivos', methods=['GET', 'POST'])
def puntos_masivos():
    if 'admin_id' not in session:
        return redirect(url_for('login'))
    rol = session.get('rol')
    if rol != 'admin':
        flash('No tienes permisos para acceder a esta página', 'error')
        return redirect(url_for('panel_admin'))

    familias = Familia.query.all()

    if request.method == 'POST':
        ids_seleccionados = request.form.getlist('familia_ids')  # lista de ids seleccionados
        tipo = request.form.get('tipo')  # 'suma' o 'canje'
        puntos = request.form.get('puntos')
        descripcion = request.form.get('descripcion')

        if not ids_seleccionados or not puntos or not tipo or not descripcion:
            flash('Completa todos los campos y selecciona al menos una familia', 'error')
            return redirect(url_for('puntos_masivos'))

        try:
            puntos = int(puntos)
            if puntos <= 0:
                flash('Los puntos deben ser un número positivo', 'error')
                return redirect(url_for('puntos_masivos'))
        except ValueError:
            flash('Los puntos deben ser un número válido', 'error')
            return redirect(url_for('puntos_masivos'))

        # Procesar cada familia seleccionada
        for fid in ids_seleccionados:
            familia = Familia.query.get(int(fid))
            if familia:
                if tipo == 'suma':
                    familia.puntos += puntos
                elif tipo == 'canje':
                    familia.puntos -= puntos
                    if familia.puntos < 0:
                        familia.puntos = 0
                # Registrar transacción
                nueva_transaccion = Transaccion(
                    familia_id=familia.id,
                    tipo=tipo,
                    puntos=puntos,
                    descripcion=descripcion,
                    fecha=datetime.now(pytz.timezone('America/Mexico_City')).replace(tzinfo=None)
                )
                db.session.add(nueva_transaccion)
                
                enviar_correo_movimiento(familia.correo, familia.nombre, puntos, tipo, descripcion)

        db.session.commit()
        
        
            # 🔢 Puntos masivos
        entry = LogEntry(
            timestamp=datetime.utcnow(),
            user=session.get("nombre_usuario"),
            role=session.get("rol"),
            action="masivo_" + tipo,
            entity="Familia",
            details=f"{len(ids_seleccionados)} familias, pts={puntos}"
        )
        db.session.add(entry)
        db.session.commit()
        
        flash(f'Se han actualizado los puntos de {len(ids_seleccionados)} familias.', 'success')
        return redirect(url_for('puntos_masivos'))

    return render_template('puntos_masivos.html', familias=familias)

from app import app, db
from flask import render_template, redirect, url_for, request, flash, session
from app.models import Familia, Transaccion, MovimientoPuntos  # ajusta nombres

# Borrar una sola familia
@app.route('/familia/<int:familia_id>/eliminar', methods=['POST'])
def eliminar_familia(familia_id):
    familia = Familia.query.get_or_404(familia_id)
    nombre_familia = familia.nombre  
    Transaccion.query.filter_by(familia_id=familia_id).delete()
    MovimientoPuntos.query.filter_by(familia_id=familia_id).delete()
    db.session.delete(familia)
    db.session.commit()

    registrar_log(
        
        "eliminar",
        "Familia",
        f"Se eliminó la familia '{nombre_familia}' (ID {familia_id})"
    )

    flash('Familia eliminada correctamente', 'success')
    return redirect(url_for('lista_familias_eliminar'))


# Borrado masivo
@app.route('/familias/eliminar_masivo', methods=['POST'])
def eliminar_familias_masivo():
    if 'admin_id' not in session:
        return redirect(url_for('login'))

    ids = request.form.getlist('familias_seleccionadas')
    if ids:
        # Borra todo de una sola pasada
        Transaccion.query.filter(Transaccion.familia_id.in_(ids)).delete(synchronize_session=False)
        MovimientoPuntos.query.filter(MovimientoPuntos.familia_id.in_(ids)).delete(synchronize_session=False)
        Familia.query.filter(Familia.id.in_(ids)).delete(synchronize_session=False)
        db.session.commit()
        
        # 📝 Registrar en LOG
        registrar_log(
            "eliminar",
            "Familia",
            f"Se eliminaron {len(ids)} familias en operación masiva"
        )

        flash(f'Se eliminaron {len(ids)} familias', 'success')
    else:
        flash('No se seleccionó ninguna familia', 'error')

    return redirect(url_for('lista_familias_eliminar'))


# Listar en la nueva plantilla
@app.route('/admin/familias/eliminar')
def lista_familias_eliminar():
    if 'admin_id' not in session:
        return redirect(url_for('login'))
    familias = Familia.query.all()
    return render_template('lista_familias_eliminar.html', familias=familias)



#LOG__________________________________________________

from flask import request, render_template
from app.models import LogEntry
from app import db
from datetime import datetime, timedelta
import pytz

from flask import request, render_template
from app.models import LogEntry
from app import db
import pytz
from datetime import datetime, timedelta

@app.route("/admin/log")
@login_requerido_admin
def log():
    # Parámetros de filtro opcionales
    usuario     = request.args.get("usuario", "")
    accion      = request.args.get("accion", "")
    entidad     = request.args.get("entidad", "")
    fecha_desde = request.args.get("desde", "")
    fecha_hasta = request.args.get("hasta", "")

    q = LogEntry.query

    if usuario:
        q = q.filter(LogEntry.user.ilike(f"%{usuario}%"))
    if accion:
        q = q.filter(LogEntry.action == accion)
    if entidad:
        q = q.filter(LogEntry.entity == entidad)

    # filtro de fechas UTC
    if fecha_desde:
        dt = datetime.strptime(fecha_desde, "%Y-%m-%d")
        q = q.filter(LogEntry.timestamp >= dt)
    if fecha_hasta:
        dt = datetime.strptime(fecha_hasta, "%Y-%m-%d") + timedelta(hours=23, minutes=59, seconds=59)
        q = q.filter(LogEntry.timestamp <= dt)

    entries = q.order_by(LogEntry.timestamp.desc()).all()

    # Convertimos UTC → hora local
    # Cambia "America/Mexico_City" por tu zona si es distinto
    local_tz = pytz.timezone("America/Mexico_City")
    for e in entries:
        # e.timestamp es datetime UTC; hacemos astimezone
        e.local_ts = e.timestamp.replace(tzinfo=pytz.utc).astimezone(local_tz)

    return render_template("log.html", entries=entries)

from datetime import datetime, timedelta, timezone

@app.before_request
def refrescar_timeout():
    if "admin_id" in session:
        # Creamos now como naive UTC
        now = datetime.utcnow()
        # Recuperamos el último timestamp (o now por defecto)
        last = session.get("last_activity", now)

        # Si last viene con tzinfo, lo convertimos a naive UTC
        if hasattr(last, "tzinfo") and last.tzinfo is not None:
            last = last.astimezone(timezone.utc).replace(tzinfo=None)

        # Comparamos
        if now - last > timedelta(minutes=30):  # tu inactividad límite
            # registrar timeout en log
            entry = LogEntry(
                user=session.get("nombre_usuario"),
                role=session.get("rol"),
                action="timeout sesión",
                entity="Admin",
                details="Cierre de sesión por inactividad"
            )
            db.session.add(entry)
            db.session.commit()
            session.clear()
            return redirect(url_for("login"))

from flask import send_file
import zipfile
import io
import os
import csv
from werkzeug.security import check_password_hash  # por si no estaba importado
from sqlalchemy import text  # ⬅️ necesario para ejecutar SQL crudo

@app.route("/resetear_base_datos", methods=["POST"])
@login_requerido_admin
def resetear_base_datos():
    if session.get("rol") != "admin":
        return "No autorizado", 403

    data = request.get_json()
    password = data.get("password", "")
    admin_id = session.get("admin_id")
    admin = Admin.query.get(admin_id)

    if not admin or not admin.verificar_password(password):
        return "Contraseña incorrecta", 401

    os.makedirs("respaldo", exist_ok=True)

    # Crear ZIP en memoria
    zip_stream = io.BytesIO()
    with zipfile.ZipFile(zip_stream, mode="w", compression=zipfile.ZIP_DEFLATED) as zipf:
        def agregar_csv(nombre, query, campos):
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(campos)
            for item in query:
                writer.writerow([getattr(item, campo) for campo in campos])
            zipf.writestr(f"{nombre}.csv", output.getvalue())

        agregar_csv("familias", Familia.query.all(), ["id", "nombre", "correo", "puntos"])
        agregar_csv("transacciones", Transaccion.query.all(), ["id", "familia_id", "tipo", "puntos", "descripcion", "fecha"])
        agregar_csv("movimientos", MovimientoPuntos.query.all(), ["id", "familia_id", "cambio", "motivo", "fecha"])
        agregar_csv("beneficios", Beneficio.query.all(), ["id", "nombre", "puntos_requeridos"])
        agregar_csv("eventos_qr", EventoQR.query.all(), ["id", "nombre_evento", "puntos", "qr_filename", "latitud", "longitud", "requiere_ubic", "valid_from", "valid_to"])
        agregar_csv("eventos_registrados", EventoQRRegistro.query.all(), ["id", "familia_id", "evento_id", "fecha"])

    zip_stream.seek(0)
    zip_filename = f"respaldo_cuponera_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"

    with open(os.path.join("respaldo", zip_filename), "wb") as f:
        f.write(zip_stream.getvalue())

    # === BORRADO + REINICIO IDS ===
    if db.engine.url.drivername.startswith("postgresql"):
        # ✅ Postgres: TRUNCATE + RESTART IDENTITY resetea los IDs a 1 automáticamente
        db.session.execute(text("""
            TRUNCATE TABLE
                evento_qr_registro,
                evento_qr,
                movimiento_puntos,
                transaccion,
                familia
            RESTART IDENTITY CASCADE;
        """))
        db.session.commit()
    else:
        # Otros motores (SQLite): borrado clásico
        db.session.query(EventoQRRegistro).delete()
        db.session.query(EventoQR).delete()
        db.session.query(MovimientoPuntos).delete()
        db.session.query(Transaccion).delete()
        db.session.query(Familia).delete()
        db.session.commit()
        # En SQLite no hace falta reiniciar secuencias; arranca en 1 tras DELETE

    # Log del reseteo (no afecta IDs de tablas que reseteamos)
    nuevo_log = LogEntry(
        timestamp=datetime.utcnow(),
        user=admin.usuario,
        role=admin.rol,
        action="eliminar",
        entity="Base de datos",
        details="El administrador ejecutó un reseteo completo de la base de datos al inicio del ciclo escolar."
    )
    db.session.add(nuevo_log)
    db.session.commit()

    return send_file(
        zip_stream,
        as_attachment=True,
        download_name=zip_filename,
        mimetype="application/zip"
    )





def enviar_correo_movimiento(destinatario, nombre_familia, puntos, tipo, motivo):
    print(f"[DEBUG] FUNCION llamada: enviar_correo_movimiento({destinatario}, {puntos}, {tipo})")
    
    from flask_mail import Message
    from app import mail
    
    if tipo == "suma":
        asunto = f"✅ Has ganado {puntos} puntos"
        cuerpo = (
            f"Hola {nombre_familia},\n\n"
            f"Has recibido {puntos} puntos por:\n👉 {motivo}\n\n"
            f"Gracias por participar en el programa Pasaporte Escolar.\n\n"
            f"-- Instituto Moderno Americano"
        )
    elif tipo == "canje":
        asunto = f"🎁 Has canjeado {puntos} puntos"
        cuerpo = (
            f"Hola {nombre_familia},\n\n"
            f"Se han descontado {puntos} puntos por:\n👉 {motivo}\n\n"
            f"¡Esperamos que disfrutes tu recompensa!\n\n"
            f"-- Instituto Moderno Americano"
        )
    else:
        asunto = "📌 Movimiento de puntos"
        cuerpo = (
            f"Hola {nombre_familia},\n\n"
            f"Se ha registrado un movimiento de puntos ({tipo}) por:\n👉 {motivo}\n\n"
            f"-- Instituto Moderno Americano"
        )

    msg = Message(subject=asunto, recipients=[destinatario], body=cuerpo)

    try:
        mail.send(msg)
        print(f"[INFO] Correo enviado a {destinatario}")
    except Exception as e:
        print(f"[ERROR] Fallo al enviar correo: {e}")
        


@app.route("/admin/editar_contrasena/<int:admin_id>", methods=["GET", "POST"])
@login_requerido_admin
def editar_contrasena_admin(admin_id):
    admin = Admin.query.get_or_404(admin_id)

    if admin.usuario == 'admin':
        flash("No se puede modificar la contraseña del administrador principal.", "error")
        return redirect(url_for("lista_administradores"))

    if request.method == "POST":
        nueva = request.form.get("nueva_contrasena")
        if not nueva or len(nueva) < 4:
            flash("La nueva contraseña debe tener al menos 4 caracteres.", "error")
        else:
            admin.set_password(nueva)
            db.session.commit()
            
            # 📝 Registrar en LOG
            registrar_log(
                "editar",
                "Admin",
                f"Se actualizó la contraseña del administrador '{admin.usuario}' (ID {admin.id})"
            )
            
            flash("Contraseña actualizada con éxito.", "success")
            return redirect(url_for("lista_administradores"))

    return render_template("editar_contrasena.html", admin=admin)


@app.route('/recuperar_contrasena', methods=['POST'])
def recuperar_contrasena():
    from flask import request, jsonify
    from flask_mail import Message
    from app import mail

    data = request.get_json() or {}
    correo = data.get('correo', '').strip()
    if not correo:
        return jsonify({'error': 'Debes proporcionar tu correo institucional.'}), 400

    familia = Familia.query.filter_by(correo=correo).first()
    if not familia:
        return jsonify({'error': 'Ese correo no está registrado.'}), 404

    # Envío de la contraseña por correo
    asunto = "Recuperación de contraseña - Pasaporte Escolar"
    cuerpo = (
        f"Hola {familia.nombre},\n\n"
        f"Tu contraseña actual es: {familia.password}\n\n"
        f"Saludos,\n"
        f"Instituto Moderno Americano"
    )
    msg = Message(subject=asunto, recipients=[correo], body=cuerpo)
    mail.send(msg)

    return jsonify({'message': 'Se ha enviado un correo con tu contraseña al correo proporcionado.'}), 200


@app.route("/status")
def status():
    return {"status": "ok"}, 200

@app.route("/eliminar_eventos_masivamente", methods=["POST"])
@login_requerido_admin
def eliminar_eventos_masivamente():
    ids = request.form.getlist("evento_ids")

    if ids:
        for evento_id in ids:
            evento = EventoQR.query.get(evento_id)
            if evento:
                # Eliminar registros hijos primero
                registros = EventoQRRegistro.query.filter_by(evento_id=evento.id).all()
                for reg in registros:
                    db.session.delete(reg)

                # Eliminar archivo QR si existe
                if evento.qr_filename:
                    qr_path = os.path.join(app.static_folder, "qr_eventos", evento.qr_filename)
                    if os.path.exists(qr_path):
                        os.remove(qr_path)

                # Eliminar el evento
                db.session.delete(evento)

        db.session.commit()
        flash(f"{len(ids)} evento(s) eliminados correctamente.", "success")
    else:
        flash("No seleccionaste ningún evento.", "error")

    return redirect(url_for("crear_qr_evento"))



from flask import send_file
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas

@app.route('/descargar_qr/<int:familia_id>')
@login_requerido_admin
def descargar_qr_admin(familia_id):
    familia = Familia.query.get_or_404(familia_id)
    qr_path = os.path.join(app.root_path, 'static', 'qr', f'familia_{familia.id}.png')

    if not os.path.exists(qr_path):
        flash("El QR no existe", "error")
        return redirect(url_for('familia', familia_id=familia.id))

    # Generar PDF temporal
    pdf_path = os.path.join(app.root_path, f"temp_qr_{familia.id}.pdf")
    c = canvas.Canvas(pdf_path, pagesize=letter)
    c.setFont("Helvetica-Bold", 18)
    c.drawCentredString(300, 750, f"QR - {familia.nombre}")
    c.drawImage(qr_path, 200, 500, width=200, height=200)
    c.showPage()
    c.save()

    return send_file(pdf_path, as_attachment=True,
                    download_name=f"QR_{familia.nombre.replace(' ', '_')}.pdf")

from textwrap import wrap

@app.route('/descargar_qr_pdf/<int:familia_id>')
@login_requerido_admin
def descargar_qr_pdf(familia_id):
    familia = Familia.query.get_or_404(familia_id)
    
    qr_path = os.path.join(app.root_path, 'static', 'qr', f'familia_{familia.id}.png')
    if not os.path.exists(qr_path):
        flash("El QR no existe", "error")
        return redirect(url_for('familia', familia_id=familia.id))

    pdf_path = os.path.join(app.root_path, f"QR_{familia.id}.pdf")
    c = canvas.Canvas(pdf_path, pagesize=letter)

    # Definir fuente y tamaño
    c.setFont("Helvetica-Bold", 20)

    # Envolver texto si es muy largo
    max_width_chars = 20  # Máximo de caracteres por línea
    nombre_familia = familia.nombre
    nombre_lineas = wrap(nombre_familia, max_width_chars)

    # Calcular posición inicial arriba del QR
    start_y = 750
    for i, linea in enumerate(nombre_lineas):
        c.drawCentredString(300, start_y - (i * 22), linea)  # 22 px de espacio entre líneas

    # Dibujar QR más abajo
    qr_y_position = start_y - (len(nombre_lineas) * 22) - 20
    c.drawImage(qr_path, 150, qr_y_position - 300, width=300, height=300)

    c.showPage()
    c.save()

    return send_file(pdf_path, as_attachment=True, download_name=f"QR_{nombre_familia.replace(' ', '_')}.pdf")

from flask import request, redirect, url_for, flash
from werkzeug.utils import secure_filename
from app import app, db
from app.models import Familia

import io, csv, tempfile, gc, os
from openpyxl import load_workbook
import qrcode

ALLOWED_EXTS = {'.xlsx', '.csv'}
BATCH_SIZE = 300  # un poco menor para memoria

def _ext(filename: str) -> str:
    idx = filename.rfind('.')
    return filename[idx:].lower() if idx != -1 else ''

def _qr_folder():
    folder = os.path.join(app.root_path, 'static', 'qr')
    os.makedirs(folder, exist_ok=True)
    return folder

def _make_family_qr(familia):
    """Genera QR con versionado como en tu flujo actual."""
    qr_url = url_for('ver_familia', familia_id=familia.id, _external=True) + f'?version_qr={familia.qr_version}'
    img = qrcode.make(qr_url)
    img.save(os.path.join(_qr_folder(), f'familia_{familia.id}.png'))

from flask import request, redirect, url_for, flash
from werkzeug.utils import secure_filename
from app import app, db
from app.models import Familia

import io, csv, tempfile, gc, os
from openpyxl import load_workbook
import qrcode
from qrcode.constants import ERROR_CORRECT_L

ALLOWED_EXTS = {'.xlsx', '.csv'}
BATCH_SIZE = 300  # tamaño de lote más seguro en Render

def _ext(filename: str) -> str:
    idx = filename.rfind('.')
    return filename[idx:].lower() if idx != -1 else ''

def _qr_folder():
    folder = os.path.join(app.root_path, 'static', 'qr')
    os.makedirs(folder, exist_ok=True)
    return folder

def _make_family_qr(familia):
    """Genera QR con versionado corto para familia"""
    qr_url = url_for('ver_familia', familia_id=familia.id) + f'?version_qr={familia.qr_version or 1}'

    qr = qrcode.QRCode(
        version=2,
        error_correction=ERROR_CORRECT_L,
        box_size=6,
        border=2
    )
    qr.add_data(qr_url)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")

    out_path = os.path.join(_qr_folder(), f'familia_{familia.id}.png')
    if os.path.exists(out_path):
        try:
            os.remove(out_path)
        except OSError:
            pass

    img.save(out_path)
    del img

from flask import request, redirect, url_for, flash, jsonify
from app import app, db
from app.models import Familia
from openpyxl import load_workbook
import io, csv, tempfile, gc, os

ALLOWED_EXTS = {'.xlsx', '.csv'}
BATCH_SIZE = 300

def _ext(filename: str) -> str:
    i = filename.rfind('.')
    return filename[i:].lower() if i != -1 else ''

@app.route("/importar_excel_familias", methods=["POST"])
@login_requerido_admin
def importar_excel_familias():
    f = request.files.get("excelFile")
    if not f or not f.filename:
        flash("❌ No se seleccionó archivo.", "error")
        return redirect(url_for("panel_admin"))

    ext = _ext(f.filename)
    if ext not in ALLOWED_EXTS:
        flash("❌ Formato no válido. Sube .xlsx o .csv", "error")
        return redirect(url_for("panel_admin"))

    try:
        insertados, duplicados = 0, 0

        if ext == ".csv":
            content = f.read()
            stream = io.TextIOWrapper(io.BytesIO(content), encoding="utf-8-sig", newline="")
            reader = csv.DictReader(stream)
            lote = []
            for row in reader:
                nombre   = (row.get("nombre") or row.get("Nombre") or "").strip()
                correo   = (row.get("correo") or row.get("Correo") or row.get("email") or "").strip().lower()
                password = (row.get("password") or row.get("Password") or "").strip()
                puntos   = row.get("puntos") or row.get("Puntos")

                if not nombre or not correo or not password:
                    continue
                if Familia.query.filter_by(correo=correo).first():
                    duplicados += 1
                    continue

                try: puntos_val = int(puntos) if puntos not in (None, "",) else 0
                except ValueError: puntos_val = 0

                fam = Familia(nombre=nombre, correo=correo, password=password, puntos=puntos_val)
                fam.qr_version = 1
                lote.append(fam)

                if len(lote) >= BATCH_SIZE:
                    db.session.add_all(lote)
                    db.session.commit()
                    insertados += len(lote)
                    lote.clear(); gc.collect()

            if lote:
                db.session.add_all(lote)
                db.session.commit()
                insertados += len(lote)
                lote.clear(); gc.collect()

        else:
            tmp_path = None
            try:
                with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                    f.save(tmp.name)
                    tmp_path = tmp.name

                wb = load_workbook(tmp_path, read_only=True, data_only=True)
                ws = wb.active

                first, headers = True, None
                lote = []
                for row in ws.iter_rows(values_only=True):
                    if first:
                        headers = [(str(c).strip().lower() if c is not None else "") for c in row]
                        first = False
                        continue

                    cells = dict(zip(headers, row))
                    nombre   = (cells.get("nombre") or "").strip() if cells.get("nombre") else ""
                    correo   = (cells.get("correo") or cells.get("email") or "").strip().lower() if (cells.get("correo") or cells.get("email")) else ""
                    password = (cells.get("password") or "").strip() if cells.get("password") else ""
                    puntos   = cells.get("puntos")

                    if not nombre or not correo or not password:
                        continue
                    if Familia.query.filter_by(correo=correo).first():
                        duplicados += 1
                        continue

                    try: puntos_val = int(puntos) if puntos not in (None, "",) else 0
                    except (ValueError, TypeError): puntos_val = 0

                    fam = Familia(nombre=nombre, correo=correo, password=password, puntos=puntos_val)
                    fam.qr_version = 1
                    lote.append(fam)

                    if len(lote) >= BATCH_SIZE:
                        db.session.add_all(lote)
                        db.session.commit()
                        insertados += len(lote)
                        lote.clear(); gc.collect()

                if lote:
                    db.session.add_all(lote)
                    db.session.commit()
                    insertados += len(lote)
                    lote.clear(); gc.collect()

                wb.close()
            finally:
                if tmp_path and os.path.exists(tmp_path):
                    try: os.remove(tmp_path)
                    except OSError: pass

        flash(f"✅ Importación completa. Insertados: {insertados}. Duplicados: {duplicados}. Ahora puedes generar los QR en lotes.", "success")
        return redirect(url_for("panel_admin"))

    except MemoryError:
        db.session.rollback()
        flash("❌ Sin memoria al importar. Divide el archivo o usa CSV.", "error")
        return redirect(url_for("panel_admin"))
    except Exception as e:
        db.session.rollback()
        flash(f"❌ Error al importar: {e}", "error")
        return redirect(url_for("panel_admin"))

from flask import jsonify, request
import qrcode, os, gc
from qrcode.constants import ERROR_CORRECT_L

def _qr_folder():
    folder = os.path.join(app.root_path, 'static', 'qr')
    os.makedirs(folder, exist_ok=True)
    return folder

def _make_family_qr_optimized(familia):
    # QR corto: ruta relativa + versionado
    from flask import url_for
    qr_url = url_for('ver_familia', familia_id=familia.id) + f'?version_qr={familia.qr_version or 1}'
    qr = qrcode.QRCode(
        version=2,
        error_correction=ERROR_CORRECT_L,
        box_size=6,
        border=2
    )
    qr.add_data(qr_url)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")

    out_path = os.path.join(_qr_folder(), f'familia_{familia.id}.png')
    if os.path.exists(out_path):
        try: os.remove(out_path)
        except OSError: pass
    img.save(out_path)
    del img

@app.route("/admin/generar_qr_pendientes", methods=["POST"])
@login_requerido_admin
def generar_qr_pendientes():
    """
    Genera QR en lotes. Parámetros opcionales en JSON:
    - last_id: último ID procesado (default 0)
    - batch_size: tamaño de lote (default 200)
    - only_missing: bool; si True, solo genera si el archivo PNG no existe (default True)
    """
    try:
        data = request.get_json(silent=True) or {}
        last_id = int(data.get("last_id", 0))
        batch_size = int(data.get("batch_size", 200))
        only_missing = bool(data.get("only_missing", True))

        familias = (Familia.query
                    .filter(Familia.id > last_id)
                    .order_by(Familia.id.asc())
                    .limit(batch_size)
                    .all())

        if not familias:
            return jsonify({"processed": 0, "last_id": last_id, "done": True})

        processed = 0
        qr_dir = _qr_folder()

        for fam in familias:
            if not getattr(fam, "qr_version", None):
                fam.qr_version = 1
                db.session.commit()

            png_path = os.path.join(qr_dir, f"familia_{fam.id}.png")
            if only_missing and os.path.exists(png_path):
                # ya existe
                pass
            else:
                _make_family_qr_optimized(fam)
            processed += 1

        gc.collect()
        new_last_id = familias[-1].id
        return jsonify({"processed": processed, "last_id": new_last_id, "done": False})

    except Exception as e:
        return jsonify({"error": str(e)}), 500


from datetime import datetime
import pytz
from flask import request, jsonify, url_for, session
from math import radians, sin, cos, sqrt, atan2
from app import db
from app.models import EventoQR, Familia, EventoQRRegistro, Transaccion

def distancia_metros(lat1, lon1, lat2, lon2):
    """Calcula distancia en metros entre dos coordenadas con Haversine"""
    R = 6371000  # radio de la Tierra en metros
    phi1, phi2 = radians(lat1), radians(lat2)
    dphi = radians(lat2 - lat1)
    dlambda = radians(lon2 - lon1)

    a = sin(dphi/2)**2 + cos(phi1) * cos(phi2) * sin(dlambda/2)**2
    c = 2 * atan2(sqrt(a), sqrt(1 - a))
    return R * c

@app.route('/validar_qr_evento', methods=["POST"])
def validar_qr_evento():
    data       = request.get_json()
    evento_id  = data.get("evento_id")
    lat        = data.get("lat")
    lon        = data.get("lon")
    familia_id = session.get("familia_id")

    # 0. Comprobación básica
    if not evento_id or not familia_id:
        return jsonify({"error": "Faltan datos"}), 400

    evento  = EventoQR.query.get(evento_id)
    familia = Familia.query.get(familia_id)
    if not evento or not familia:
        return jsonify({"error": "Evento o familia no encontrada"}), 404

    # 1. Fecha/hora actual en zona local
    tz          = pytz.timezone("America/Mexico_City")
    ahora_local = datetime.now(tz)

    # valid_from
    if evento.valid_from:
        vf_utc   = evento.valid_from.replace(tzinfo=pytz.UTC)
        vf_local = vf_utc.astimezone(tz)
        if ahora_local < vf_local:
            inicio = vf_local.strftime('%d/%m/%Y %H:%M')
            return jsonify({
                "error": f"El evento aún no está activo. Desde {inicio}",
                "code": "fuera_de_fecha"
            }), 400

    # valid_to
    if evento.valid_to:
        vt_utc   = evento.valid_to.replace(tzinfo=pytz.UTC)
        vt_local = vt_utc.astimezone(tz)
        if ahora_local > vt_local:
            fin = vt_local.strftime('%d/%m/%Y %H:%M')
            return jsonify({
                "error": f"El evento expiró. Hasta {fin}",
                "code": "fuera_de_fecha"
            }), 400

    # 2. Validación de ubicación (solo si requiere_ubic=True)
    if evento.requiere_ubic:
        if lat is None or lon is None:
            return jsonify({"error": "No se proporcionaron coordenadas"}), 400

        distancia_m = distancia_metros(
            float(evento.latitud), float(evento.longitud),
            float(lat), float(lon)
        )
        if distancia_m > 500:  # mismo límite que ya usabas
            return jsonify({
                "redirect": url_for("ubicacion_invalida", evento_id=evento.id)
            })

    # 3. Checar duplicados
    if EventoQRRegistro.query.filter_by(
        evento_id=evento.id,
        familia_id=familia.id
    ).first():
        return jsonify({
            "redirect": url_for(
                "asistencia_exitosa",
                evento_id=evento.id,
                ya_asistio=1
            )
        })

    # 4. Registrar asistencia y sumar puntos
    familia.puntos += evento.puntos
    db.session.add(EventoQRRegistro(
        familia_id=familia.id,
        evento_id=evento.id
    ))
    db.session.add(Transaccion(
        familia_id=familia.id,
        tipo='suma',
        puntos=evento.puntos,
        descripcion=f"Asistencia al evento: {evento.nombre_evento}"
    ))

    db.session.commit()
    
    # 📝 Registrar en LOG
    registrar_log(
        "escanear",
        "EventoQR",
        f"Familia '{familia.nombre}' (ID {familia.id}) escaneó el evento '{evento.nombre_evento}' (ID {evento.id})"
    )

    # 💌 Enviar correo de confirmación
    enviar_correo_movimiento(
        destinatario=familia.correo,
        tipo='suma',
        puntos=evento.puntos,
        nombre_familia=familia.nombre,
        motivo=f"Asistencia al evento: {evento.nombre_evento}"
    )

    return jsonify({
        "redirect": url_for("asistencia_exitosa", evento_id=evento.id)
    })


